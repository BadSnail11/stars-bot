# src/handlers/stats_excel.py
from aiogram import Router, types, F
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from sqlalchemy import select, func, or_
from sqlalchemy.sql import nulls_last
from typing import List, Optional, Tuple
import io
from datetime import datetime, timedelta, timezone
import re

from src.db import SessionLocal
from src.models import User, Order, Withdrawal  # имя класса — как в твоих моделях

from src.utils.owner_scope import resolve_owner_and_bot_key
from ..keyboards.common import nav_to_menu, stats_root_kb, withdrawals_filter_kb



# XLSX
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

router = Router(name="stats_excel")

# ----------------- FSM -----------------
class StatsStates(StatesGroup):
    idle = State()
    waiting_user_query = State()
    waiting_user_filter = State()
    waiting_period_start = State()      # ожидание начальной даты
    waiting_period_end = State()        # ожидание конечной даты

# ----------------- Date Helpers -----------------
def _now_utc() -> datetime:
    return datetime.now(timezone.utc)

def parse_date(date_str: str) -> Optional[datetime]:
    """Парсит дату в форматах: DD.MM.YYYY, DD-MM-YYYY, YYYY-MM-DD"""
    date_str = date_str.strip()
    
    # Попробовать DD.MM.YYYY
    match = re.match(r'^(\d{1,2})\.(\d{1,2})\.(\d{4})$', date_str)
    if match:
        day, month, year = map(int, match.groups())
        try:
            return datetime(year, month, day, tzinfo=timezone.utc)
        except ValueError:
            return None
    
    # Попробовать DD-MM-YYYY
    match = re.match(r'^(\d{1,2})-(\d{1,2})-(\d{4})$', date_str)
    if match:
        day, month, year = map(int, match.groups())
        try:
            return datetime(year, month, day, tzinfo=timezone.utc)
        except ValueError:
            return None
    
    # Попробовать YYYY-MM-DD
    match = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', date_str)
    if match:
        year, month, day = map(int, match.groups())
        try:
            return datetime(year, month, day, tzinfo=timezone.utc)
        except ValueError:
            return None
    
    return None

def format_date(dt: datetime) -> str:
    """Форматирует дату для отображения"""
    return dt.strftime("%d.%m.%Y")

def get_period_keyboard(cancel_prefix: str = "st_home"):
    """Клавиатура для отмены ввода даты"""
    return types.InlineKeyboardMarkup(inline_keyboard=[
        [types.InlineKeyboardButton(text="❌ Отменить", callback_data=cancel_prefix)]
    ])

# ----------------- Root -----------------

@router.callback_query(F.data == "stats")
async def stats_root(cb: types.CallbackQuery, state: FSMContext):
    await state.set_state(StatsStates.idle)
    await cb.message.edit_text("📊 Выберите, что экспортировать:", reply_markup=stats_root_kb())
    await cb.answer()

@router.callback_query(F.data == "statistics")
async def st_home(cb: types.CallbackQuery, state: FSMContext):
    await stats_root(cb, state)

# ----------------- XLSX utils -----------------
def _auto_width(ws):
    widths = {}
    for row in ws.rows:
        for cell in row:
            val = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), len(val))
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(w + 2, 10), 60)

def _wb_from_table(sheet_title: str, headers: List[str], rows: List[Tuple]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Sheet1"
    ws.append(headers)
    for c in ws[1]:
        c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(list(r))
    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

async def _bot_id_by_admin(s, admin_tg_id: int) -> Optional[int]:
    _, bot_id = await resolve_owner_and_bot_key(s, admin_tg_id)
    return bot_id

# ----------------- 1) Все заказы -----------------
def orders_filter_kb(scope: str, user_id: Optional[int] = None):
    uid = f"_{user_id}" if (scope == "user" and user_id is not None) else ""
    return types.InlineKeyboardMarkup(inline_keyboard=[
        [
            types.InlineKeyboardButton(text="✅ Только успешные", callback_data=f"st_do_{scope}_paid{uid}"),
            types.InlineKeyboardButton(text="Все", callback_data=f"st_do_{scope}_all{uid}")
        ],
        [types.InlineKeyboardButton(text="⬅️ В разделы", callback_data="st_home")]
    ])

@router.callback_query(F.data == "st_orders")
async def st_orders(cb: types.CallbackQuery, state: FSMContext):
    await state.set_state(StatsStates.idle)
    await cb.message.edit_text("Экспорт заказов бота. Сначала выберите фильтр:", reply_markup=orders_filter_kb(scope="all"))
    await cb.answer()

# — выбрали paid/all → запрашиваем начальную дату
@router.callback_query(F.data.in_(("st_do_all_paid", "st_do_all_all")))
async def st_orders_pick_period(cb: types.CallbackQuery, state: FSMContext):
    key = "st_do_all_paid" if cb.data.endswith("_paid") else "st_do_all_all"
    await state.update_data(mode=key)
    await state.set_state(StatsStates.waiting_period_start)
    
    text = (
        "📅 Введите начальную дату периода (включительно)\n"
        "Форматы: DD.MM.YYYY, DD-MM-YYYY, YYYY-MM-DD\n"
        "Пример: 01.12.2024"
    )
    await cb.message.edit_text(text, reply_markup=get_period_keyboard("st_home"))
    await cb.answer()

# — ввод начальной даты
@router.message(StatsStates.waiting_period_start)
async def process_start_date(message: types.Message, state: FSMContext):
    date = parse_date(message.text)
    if not date:
        await message.answer(
            "❌ Неверный формат даты. Используйте: DD.MM.YYYY, DD-MM-YYYY или YYYY-MM-DD\n"
            "Попробуйте снова:",
            reply_markup=get_period_keyboard("st_home")
        )
        return
    
    await state.update_data(start_date=date)
    await state.set_state(StatsStates.waiting_period_end)
    
    await message.answer(
        f"✅ Начальная дата: {format_date(date)}\n"
        "📅 Теперь введите конечную дату периода (включительно):",
        reply_markup=get_period_keyboard("st_home")
    )

# — ввод конечной даты
@router.message(StatsStates.waiting_period_end)
async def process_end_date(message: types.Message, state: FSMContext):
    end_date = parse_date(message.text)
    if not end_date:
        await message.answer(
            "❌ Неверный формат даты. Используйте: DD.MM.YYYY, DD-MM-YYYY или YYYY-MM-DD\n"
            "Попробуйте снова:",
            reply_markup=get_period_keyboard("st_home")
        )
        return
    
    data = await state.get_data()
    start_date = data['start_date']
    
    # Проверка, что конечная дата не раньше начальной
    if end_date < start_date:
        await message.answer(
            "❌ Конечная дата не может быть раньше начальной.\n"
            "Введите конечную дату снова:",
            reply_markup=get_period_keyboard("st_home")
        )
        return
    
    # Добавляем время 23:59:59 к конечной дате
    end_date = end_date.replace(hour=23, minute=59, second=59)
    
    await state.update_data(end_date=end_date)
    
    # Определяем тип отчета и генерируем его
    data = await state.get_data()
    mode = data.get('mode', '')
    
    if mode == 'st_users':
        await generate_users_report(message, state)
    elif mode.startswith('st_do_user_'):
        await generate_user_orders_report(message, state)
    elif mode.startswith('st_do_withdrawals_'):          # ← НОВОЕ
        await generate_withdrawals_report(message, state)
    else:
        await generate_orders_report(message, state)

async def generate_orders_report(message: types.Message, state: FSMContext):
    data = await state.get_data()
    mode = data['mode']
    start_date = data['start_date']
    end_date = data['end_date']
    
    only_paid = (mode == "st_do_all_paid")
    
    async with SessionLocal() as s:
        bot_id = await _bot_id_by_admin(s, message.chat.id)
        if not bot_id:
            await message.answer("Зеркальный бот не найден.", reply_markup=nav_to_menu())
            await state.clear()
            return

        q = select(
            Order.id.label("order_id"),
            User.tg_user_id, User.username, User.first_name, User.last_name,
            Order.type, Order.amount, Order.price, Order.income, Order.currency,
            Order.status, Order.recipient, Order.created_at, Order.paid_at
        ).join(User, User.id == Order.user_id).where(User.bot_id == bot_id)

        # Преобразуем даты в UTC naive для сравнения с базой данных
        start_date_naive = start_date.replace(tzinfo=None)
        end_date_naive = end_date.replace(tzinfo=None)

        if only_paid:
            q = q.where(Order.status == "paid", Order.paid_at >= start_date_naive, Order.paid_at <= end_date_naive)
        else:
            q = q.where(Order.created_at >= start_date_naive, Order.created_at <= end_date_naive)

        q = q.order_by(
            Order.paid_at.desc().nullslast(),
            Order.created_at.desc()
        )
        rows = (await s.execute(q)).all()

    headers = ["OrderID","TG UserID","Username","First Name","Last Name",
               "Type","Amount","Price","Income","Currency","Status","Recipient",
               "CreatedAt","PaidAt"]
    data_rows = []
    for r in rows:
        cr = r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else ""
        pr = r.paid_at.strftime("%Y-%m-%d %H:%M:%S") if r.paid_at else ""
        data_rows.append((
            r.order_id, r.tg_user_id, r.username, r.first_name, r.last_name,
            r.type, r.amount, float(r.price or 0), float(r.income or 0), r.currency,
            r.status, r.recipient, cr, pr
        ))

    period_str = f"{start_date.strftime('%d%m%Y')}_{end_date.strftime('%d%m%Y')}"
    mode_str = "paid" if only_paid else "all"
    
    xbytes = _wb_from_table(
        sheet_title=f"orders_{mode_str}",
        headers=headers, rows=data_rows
    )
    fname = f"orders_{mode_str}_{period_str}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    await message.bot.send_document(
        chat_id=message.chat.id,
        document=types.BufferedInputFile(xbytes, filename=fname),
        caption=f"📊 Отчет за период: {format_date(start_date)} - {format_date(end_date)}"
    )
    await state.clear()

# ----------------- 2) Пользователи бота -----------------
@router.callback_query(F.data == "st_users")
async def st_users_pick_period(cb: types.CallbackQuery, state: FSMContext):
    await state.update_data(mode="st_users")
    await state.set_state(StatsStates.waiting_period_start)
    
    text = (
        "📅 Введите начальную дату периода для экспорта пользователей (включительно)\n"
        "Форматы: DD.MM.YYYY, DD-MM-YYYY, YYYY-MM-DD\n"
        "Пример: 01.12.2024"
    )
    await cb.message.edit_text(text, reply_markup=get_period_keyboard("st_home"))
    await cb.answer()

async def generate_users_report(message: types.Message, state: FSMContext):
    data = await state.get_data()
    start_date = data['start_date']
    end_date = data['end_date']
    
    # Преобразуем даты в UTC naive для сравнения с базой данных
    start_date_naive = start_date.replace(tzinfo=None)
    end_date_naive = end_date.replace(tzinfo=None)
    
    async with SessionLocal() as s:
        bot_id = await _bot_id_by_admin(s, message.chat.id)
        if not bot_id:
            await message.answer("Зеркальный бот не найден.", reply_markup=nav_to_menu())
            await state.clear()
            return

        q = select(
            User.id, User.tg_user_id, User.username, User.first_name, User.last_name,
            User.lang_code, User.balance, User.accepted_offer_at, User.is_blocked, User.created_at
        ).where(User.bot_id == bot_id, User.created_at >= start_date_naive, User.created_at <= end_date_naive).order_by(User.created_at.desc())
        rows = (await s.execute(q)).all()

    headers = ["ID","TG UserID","Username","First Name","Last Name","Lang",
               "Balance","AcceptedOfferAt","IsBlocked","CreatedAt"]
    data_rows = []
    for r in rows:
        ao = r.accepted_offer_at.strftime("%Y-%m-%d %H:%M:%S") if r.accepted_offer_at else ""
        cr = r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else ""
        data_rows.append((
            r.id, r.tg_user_id, r.username, r.first_name, r.last_name,
            r.lang_code, float(r.balance or 0), ao, bool(r.is_blocked), cr
        ))

    period_str = f"{start_date.strftime('%d%m%Y')}_{end_date.strftime('%d%m%Y')}"
    
    xbytes = _wb_from_table(f"users_{period_str}", headers, data_rows)
    fname = f"users_{period_str}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    await message.bot.send_document(
        chat_id=message.chat.id,
        document=types.BufferedInputFile(xbytes, filename=fname),
        caption=f"👥 Пользователи за период: {format_date(start_date)} - {format_date(end_date)}"
    )
    await state.clear()

# ----------------- 3) Заказы конкретного пользователя -----------------
def ask_user_kb():
    return types.InlineKeyboardMarkup(inline_keyboard=[
        [types.InlineKeyboardButton(text="⬅️ В разделы", callback_data="st_home")],
    ])

@router.callback_query(F.data == "st_user_orders")
async def st_user_orders(cb: types.CallbackQuery, state: FSMContext):
    m = cb.message
    await state.set_state(StatsStates.waiting_user_query)
    await m.edit_text("Введите @username или числовой tg_user_id пользователя:", reply_markup=ask_user_kb())
    await cb.answer()

@router.message(StatsStates.waiting_user_query)
async def st_user_pick(m: types.Message, state: FSMContext):
    q = (m.text or "").strip()
    if not q:
        await m.answer("Пустой ввод. Повторите.", reply_markup=nav_to_menu()); return

    async with SessionLocal() as s:
        bot_id = await _bot_id_by_admin(s, m.chat.id)
        if not bot_id:
            await m.answer("Зеркальный бот не найден.", reply_markup=nav_to_menu()); return

        conds = []
        if q.startswith("@"):
            uname = q[1:].lower()
            conds.append(func.lower(User.username) == uname)
        elif q.isdigit():
            conds.append(User.tg_user_id == int(q))
        else:
            await m.answer("Ожидал @username или числовой tg_user_id.", reply_markup=nav_to_menu()); return

        res = await s.execute(
            select(User.id, User.username, User.tg_user_id)
            .where(User.bot_id == bot_id)
            .where(or_(*conds)).limit(1)
        )
        row = res.first()

    if not row:
        await m.answer("Пользователь не найден.", reply_markup=nav_to_menu()); return

    user_id, username, tg_user_id = row
    await state.update_data(bot_id=bot_id, user_id=user_id, username=(username or ""), tg=tg_user_id)

    # спрашиваем фильтр paid/all
    await m.answer(
        f"Пользователь: {('@' + username) if username else f'id{tg_user_id}'}\nВыберите фильтр:",
        reply_markup=orders_filter_kb(scope="user", user_id=user_id)
    )
    await state.set_state(StatsStates.waiting_user_filter)

# выбрали paid/all → запрашиваем начальную дату
@router.callback_query(StatsStates.waiting_user_filter, F.data.regexp(r"^st_do_user_(paid|all)_(\d+)$"))
async def st_user_pick_period(cb: types.CallbackQuery, state: FSMContext):
    parts = cb.data.split("_")
    mode, uid = parts[2], parts[3]
    await state.update_data(user_mode=mode, user_uid=int(uid), mode=f"st_do_user_{mode}_{uid}")
    await state.set_state(StatsStates.waiting_period_start)
    
    text = (
        "📅 Введите начальную дату периода (включительно)\n"
        "Форматы: DD.MM.YYYY, DD-MM-YYYY, YYYY-MM-DD\n"
        "Пример: 01.12.2024"
    )
    await cb.message.edit_text(text, reply_markup=get_period_keyboard("st_home"))
    await cb.answer()


@router.callback_query(F.data == "st_withdrawals")
async def st_withdrawals(cb: types.CallbackQuery, state: FSMContext):
    await state.set_state(StatsStates.idle)
    await cb.message.edit_text("Экспорт выводов. Сначала выберите фильтр:", reply_markup=withdrawals_filter_kb())
    await cb.answer()

# выбрали paid/all → запрашиваем начальную дату
@router.callback_query(F.data.in_(("st_do_withdrawals_paid", "st_do_withdrawals_all")))
async def st_withdrawals_pick_period(cb: types.CallbackQuery, state: FSMContext):
    key = "st_do_withdrawals_paid" if cb.data.endswith("_paid") else "st_do_withdrawals_all"
    await state.update_data(mode=key)
    await state.set_state(StatsStates.waiting_period_start)
    text = (
        "📅 Введите начальную дату периода (включительно)\n"
        "Форматы: DD.MM.YYYY, DD-MM-YYYY, YYYY-MM-DD\n"
        "Пример: 01.12.2024"
    )
    await cb.message.edit_text(text, reply_markup=get_period_keyboard("st_home"))
    await cb.answer()


# Генерация отчетов для заказов конкретного пользователя
async def generate_user_orders_report(message: types.Message, state: FSMContext):
    data = await state.get_data()
    start_date = data['start_date']
    end_date = data['end_date']
    user_mode = data['user_mode']
    user_id = data['user_uid']
    
    only_paid = (user_mode == "paid")
    
    # Преобразуем даты в UTC naive для сравнения с базой данных
    start_date_naive = start_date.replace(tzinfo=None)
    end_date_naive = end_date.replace(tzinfo=None)

    async with SessionLocal() as s:
        q = select(
            Order.id.label("order_id"),
            Order.type, Order.amount, Order.price, Order.income, Order.currency,
            Order.status, Order.recipient, Order.created_at, Order.paid_at
        ).where(Order.user_id == user_id)

        if only_paid:
            q = q.where(Order.status == "paid", Order.paid_at >= start_date_naive, Order.paid_at <= end_date_naive)
        else:
            q = q.where(Order.created_at >= start_date_naive, Order.created_at <= end_date_naive)

        q = q.order_by(
            Order.paid_at.desc().nullslast(),
            Order.created_at.desc()
        )
        rows = (await s.execute(q)).all()

        u = (await s.execute(select(User).where(User.id == user_id))).scalar_one_or_none()

    headers = ["OrderID","Type","Amount","Price","Income","Currency",
               "Status","Recipient","CreatedAt","PaidAt"]
    data_rows = []
    for r in rows:
        cr = r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else ""
        pr = r.paid_at.strftime("%Y-%m-%d %H:%M:%S") if r.paid_at else ""
        data_rows.append((
            r.order_id, r.type, r.amount, float(r.price or 0),
            float(r.income or 0), r.currency, r.status, r.recipient, cr, pr
        ))

    uname = (f"@{u.username}" if (u and u.username) else f"id{u.tg_user_id if u else user_id}")
    period_str = f"{start_date.strftime('%d%m%Y')}_{end_date.strftime('%d%m%Y')}"
    
    xbytes = _wb_from_table(
        sheet_title=f"user_orders_{user_mode}",
        headers=headers, rows=data_rows
    )
    fname = f"user_orders_{uname.replace('@','')}_{user_mode}_{period_str}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    await message.bot.send_document(
        chat_id=message.chat.id,
        document=types.BufferedInputFile(xbytes, filename=fname),
        caption=f"📊 Заказы пользователя {uname} за период: {format_date(start_date)} - {format_date(end_date)}"
    )
    await state.clear()


async def generate_withdrawals_report(message: types.Message, state: FSMContext):
    data = await state.get_data()
    mode = data['mode']  # 'st_do_withdrawals_paid' | 'st_do_withdrawals_all'
    start_date = data['start_date']
    end_date = data['end_date']

    only_success = (mode == "st_do_withdrawals_paid")
    # naive UTC для TIMESTAMP WITHOUT TIME ZONE
    start_date_naive = start_date.replace(tzinfo=None)
    end_date_naive = end_date.replace(tzinfo=None)

    # статусы успеха — «sent» (по нашей договорённости)
    SUCCESS_STATUSES = {"sent", "completed", "success"}

    async with SessionLocal() as s:
        bot_id = await _bot_id_by_admin(s, message.chat.id)
        if not bot_id:
            await message.answer("Зеркальный бот не найден.", reply_markup=nav_to_menu())
            await state.clear()
            return

        # join withdrawals -> users, фильтр по users.bot_id
        q = select(
            Withdrawal.id.label("withdrawal_id"),
            User.tg_user_id, User.username, User.first_name, User.last_name,
            Withdrawal.amount, Withdrawal.currency,
            Withdrawal.to_address, Withdrawal.status,
            Withdrawal.created_at, Withdrawal.processed_at
        ).join(User, User.id == Withdrawal.user_id).where(User.bot_id == bot_id)

        if only_success:
            # по успешным — период по processed_at
            q = q.where(
                Withdrawal.status.in_(SUCCESS_STATUSES),
                Withdrawal.processed_at >= start_date_naive,
                Withdrawal.processed_at <= end_date_naive
            )
        else:
            # по всем — период по created_at
            q = q.where(
                Withdrawal.created_at >= start_date_naive,
                Withdrawal.created_at <= end_date_naive
            )

        q = q.order_by(
            Withdrawal.processed_at.desc().nullslast(),
            Withdrawal.created_at.desc()
        )
        rows = (await s.execute(q)).all()

    headers = [
        "WithdrawalID","TG UserID","Username","First Name","Last Name",
        "AmountTON","Currency","ToAddress","Status","ProviderID",
        "CreatedAt","ProcessedAt"
    ]

    data_rows = []
    for r in rows:
        cr = r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else ""
        pr = r.processed_at.strftime("%Y-%m-%d %H:%M:%S") if r.processed_at else ""
        data_rows.append((
            r.withdrawal_id, r.tg_user_id, r.username, r.first_name, r.last_name,
            float(r.amount or 0),
            r.currency, r.to_address, r.status, r.provider_id or "",
            cr, pr
        ))

    period_str = f"{start_date.strftime('%d%m%Y')}_{end_date.strftime('%d%m%Y')}"
    mode_str = "success" if only_success else "all"

    xbytes = _wb_from_table(
        sheet_title=f"withdrawals_{mode_str}",
        headers=headers, rows=data_rows
    )
    fname = f"withdrawals_{mode_str}_{period_str}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    await message.bot.send_document(
        chat_id=message.chat.id,
        document=types.BufferedInputFile(xbytes, filename=fname),
        caption=f"💸 Выводы за период: {format_date(start_date)} - {format_date(end_date)}"
    )
    await state.clear()
